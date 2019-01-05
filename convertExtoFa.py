import openpyxl

wb = openpyxl.load_workbook('DNARNAPROT.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
'''
print(wb.get_sheet_names())
print(sheet['B8'].value)
print(sheet.max_row)
'''
for i in range(1, sheet.max_row + 1):
l = r"C:\Users\dell\Desktop\celldata" + str(i) + ".txt"
f = open(l, 'w+')
print("OPENED FILE: celldata"+str(i)+".txt")
stri = 'A' + str(i) # fetching 1st column data--assume seq id
strj = 'C' + str(i) # fetching 2nd column data--assume sequence
phospholipid = sheet[strj].value
f.write('>' + sheet[stri].value + '\n')
if len(phospholipid) < 60:
f.write(phospholipid)
else:
while len(phospholipid) > 60:
f.write(phospholipid[0:59] + '\n')
phospholipid = phospholipid[60:]
f.write(phospholipid)
